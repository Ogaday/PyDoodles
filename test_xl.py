# http://stackoverflow.com/questions/34205998/python-openpyxl-copy-values-from-column-a-to-b-with-same-rows-range/

import openpyxl

wb = openpyxl.load_workbook('sheet.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

translated = []
for row in sheet.iter_rows('A2:A74'):
    for cell in row:
        translated.append(cell.value)

for row, val in zip(sheet.iter_rows('B2:B74'), translated):
    for cell in row:
        cell.value = val

wb.save("sheet.xlsx")